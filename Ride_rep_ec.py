# IMMPORT LIBRARIES_____________________________________________________________________________________________________
from math import sqrt
import openpyxl as oxl
import googlemaps
from googlemaps import Client
from datetime import datetime
import unicodedata
from geopy.distance import great_circle
import copy
import numpy as np
import folium
from folium import plugins
import tkinter
from tkinter import *
import bs4
from bs4 import BeautifulSoup
import functools
from functools import partial 
from pyroutelib3 import Router
from IPython.display import display
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import base64
import webbrowser

# DECLARE GLOBAL VARIABLES AND CONSTANT_________________________________________________________________________________

# Definir les attributs d'un salarié
# salarie est la liste des salarié. Pour acceder à un salarié en particulier dans cette liste, on indiquera salarie[i]
# pour acceder aux attributs de ce salarie, on navigue dans le dictionnaire en indiquant salarie[i][attribut][sous-attribut]
salarie = []
attributsSalarie = {'adresse': {'domicile': None, 'affectation': None},
                        'geolocation': {'domicile': None, 'affectation': None},
                        'distance': {'voiture': 'Adresse Invalide', 'velo': None},
                        'temps': {'voiture': None, 'transport': None, 'velo': 0, 'VAE': None, 'textVoiture': None, 'textTransport': None},
                        'CO2': {'voiture': None, 'transport': 'Indisponible', 'velo': None, 'VAE': None},
                        'etapes': {'transport': None, 'velo': None, 'voiture': None},
                        'path': {'voiture': [], 'transport': [], 'velo': []},
                        'elevation': {'velo': None},
                        'partenaires': {'coworkers': None, 'voisins': None},
                        'solution RIDE': {'voiture': None, 'transport': None, 'velo': None, 'VAE': None},
                        'info':{'nom': None, 'encadrant': None, 'emploi': None, 'ecart-type':0}} #Ajout de l'écart-type des distances chantiers
chantiers = []
attributsChantier = {'distance': {'voiture': [], 'velo': []},
                     'temps': {'voiture': [], 'transport': [], 'velo': [], 'VAE': [], 'textVoiture': [], 'textTransport': []},   
                    'personnes' : [],
                    'geolocation': None,
                    'adresse': None,
                    'choix':[],
                    'places':1,
                    'plein': False,
                    'path':{'voiture': [], 'transport': [], 'velo': []},
                    'etapes': {'voiture':[],'transport': [], 'velo': []},
                    }

# Google
# L'API Google Direction permet de récupérer les informations relatives à un trajet (exactement comme google maps)
# D'autres API sont disponibles telles que Elevation, ou Geocoding (voir à ce sujet : https://console.cloud.google.com/apis/library/)
# Pour avoir l'autorisation d'utiliser cette API (c'est à dire récupérer des information dans la base de données de google), il faut une clé (API_KEY)
# Une clé peut être créé à partir de n'importe quel compte Google. Pour une clé gratuite, la limite de demande est de 2500/jour.

API_KEY = 'AIzaSyCNzx3q_glPVMjmYQwgr9kcH4QIeeqShx8'

gmaps = Client(key=API_KEY)

    
# Excel
# (On indique la plage (en vertical) du tableau excel à traiter)
begin = 48 # 3
maxRow = 54  #117 # int() # 44 à la base 

# Vitesses moyennes 
vitesseMoyenneVelo = 15  # km/h
vitesseMoyenneVAE = 19  # km/h

# CO2 par types de transport
CO2VoitureEssence = 190  # g/km
CO2VoitureDiesel = 160  # g/km
CO2VoitureElectrique = 80  # g/km
CO2Velo = 5  # g/km
CO2VAE = 16  # g/km
CO2Rer = 5.7  # g/km
CO2Tram = 6  # g/km
CO2Metro = 5.7  # g/km
CO2Bus = 154  # g/km


# criteresMax
tempsVeloMax = 20  # min
tempsVAEMax = 20  # min
tempsTransportMax = 35  # min
distNeighborsMax = 0.5  # km


# création des variables
listeIndexaTraiter = []
listeError = []
#isteChant = []  # Liste des chantiers associées au checkbutton

colorSet = {'white' : '#FFFFFF', 'green' : '#71AE26', 'red' : '#D33D2A', 'blue' : '#38A9DB', 'black' : '#000000', 'purple' : '#725394', 'orange':'#ff9900'}

# FUNCTIONS_____________________________________________________________________________________________________________


def init_chantier(): 
    
    chantier_dans_liste = False
    
    for item in salarie :  # Parcours les 2 listes, si la fonction trouve un chantier dans la liste salarie mais pas dans la liste chantiers, elle l'ajoute
        
        chantier_dans_liste = False
        
        for chant in chantiers :
            if item['adresse']['affectation'] == chant['adresse']:
                
                chantier_dans_liste = True
                chant['places'] += 1
                continue # Sortie de boucle si le chantier est déjà dans la liste des chantiers
        
        if chantier_dans_liste == False:
            
            chantiers.append({'temps': {'voiture': [], 'transport': [], 'velo': [], 'VAE': [], 'textVoiture': [], 'textTransport': []},'etapes': {'voiture':[], 'transport': [], 'velo': []}, 'distance':{'velo' :[], 'voiture':[]}, 'choix': [],'adresse' :item['adresse']['affectation'],'places':1,'plein': False, 'path': {'voiture': [], 'transport': [], 'velo': []}, 'geolocation':None}) 
        
def getGeo(): 
    
    for chant in chantiers:

            geocode_afc = gmaps.geocode(address = chant['adresse']) 
        
            try :
             # Affectation des latitude et Longitude des adresses chantier
             chant['geolocation'] =(geocode_afc[0]["geometry"]["location"]["lat"],geocode_afc[0]["geometry"]["location"]["lng"])
             
             
             
            except : continue    
        
def getItineraire(liste):
    """
    Recupérer les informations de l'itinéraire calculé par Google Directions API (gmaps.directions())
    """
    now = datetime.now()
    gmaps = googlemaps.Client(API_KEY)
    myMode = ['driving','transit', 'bicycling']#'driving', 'transit', 'bicycling'] 
    
    #now = datetime(2019, 7, 29, 7, 0, 0)   # datetime (year, month, day[, hour[, minute[, second[, microsecond[, tzinfo]]]]])
    
    for i in range(L):  # pour chaque salarie
        avancement = round(((i - 0) / (L - 0)) * (100 - 0) + 0, 2)
        
        for j in range(0, len(myMode)):  # et pour chaque mode de transport

            #try:  # essaye de trouver un itinéraire entre domicile et affectation pour le mode j
                # Les information de Google Directions API se trouvent à cette adresse
                # https://developers.google.com/maps/documentation/directions/intro?hl=fr#TravelModes
                # Les arguments principaux que nous utilisons sont : adresse de domicile, adresse d'affectation, mode, region de recherche d'adresse, la date et l'heure de départ, la langue des résultats, le type de traffic pris en compte)
                # Directions API retourne un dictionnaire de résultats : voir la structure du dictionnaire au chapitre "example de résultat" de la page cité ci-dessus
        
            #if salarie[i]['adresse']['affectation'] in listeChant: # Si le salarie se trouve dans les chantiers cochés
        
               for chant in chantiers:
   
                    myItineraire = gmaps.directions(liste[i]['adresse']['domicile'], chant['adresse'],mode=str(myMode[j]), region='FR', departure_time=now, language='fr', traffic_model='best_guess')
                    
                    print(i)
                    print(j)
                    if myItineraire :  # si myItineraire existe (c'est à dire si google à trouver des résultats)
                        liste[i]['geolocation']['domicile'] = (myItineraire[0]['legs'][0]['start_location']['lat'], myItineraire[0]['legs'][0]['start_location']['lng'])  # tu ajoutes un tupple (lat, lng) dans le dictionnaire aux clés [géolocation][domicile] que l'on trouve dans le dictionnaire de résultats de google
                        liste[i]['geolocation']['affectation'] = (myItineraire[0]['legs'][0]['end_location']['lat'],myItineraire[0]['legs'][0]['end_location']['lng'])  # idem pour affectation
                        
                    
                        if myMode[j] == 'driving':  # quand j est 'driving'
                        
                           
                            chant['distance']['voiture'].append(formatValueNum(myItineraire[0]['legs'][0]['distance']['value'],'km') )
                            chant['choix'].append(False)
                            chant['path']['voiture'].append(myItineraire[0]['overview_polyline']['points'])
                            chant['temps']['voiture'].append(myItineraire[0]['legs'][0]['duration_in_traffic']['value']*0.016666667)
                            chant['etapes']['voiture'].append(myItineraire[0]['legs'][0]['steps'])
                            
                            if liste[i]['adresse']['affectation'] == chant['adresse']:  #On garde ces instructions pour pouvoir établir la moyenne de temps de voiture sans répartition
                                liste[i]['distance']['voiture'] = formatValueNum(myItineraire[0]['legs'][0]['distance']['value'],'km')  # tu vas chercher le résultat de distance en voiture que tu formate en km, et tu l'ajoute aux attributs du salarié
                                liste[i]['temps']['voiture'] = myItineraire[0]['legs'][0]['duration_in_traffic']['value']*0.016666667  # idem pour le temps
                                liste[i]['temps']['textVoiture'] = myItineraire[0]['legs'][0]['duration_in_traffic']['text']
                                liste[i]['path']['voiture'] = myItineraire[0]['overview_polyline']['points']  # idem pour les étapes
                                liste[i]['etapes']['voiture'] = myItineraire[0]['legs'][0]['steps'] 
                        
                        if myMode[j] == 'transit':  # quand j est 'transit'
                            
                                chant['temps']['transport'].append(myItineraire[0]['legs'][0]['duration']['value']*0.016666667)
                                chant['path']['transport'].append(myItineraire[0]['overview_polyline']['points'])
                                chant['etapes']['transport'].append(myItineraire[0]['legs'][0]['steps'])
                                chant['temps']['textTransport'].append(myItineraire[0]['legs'][0]['duration']['text'])
                                
                                liste[i]['temps']['transport'] = myItineraire[0]['legs'][0]['duration']['value']*0.016666667
                                liste[i]['etapes']['transport'] = myItineraire[0]['legs'][0]['steps']
                                liste[i]['path']['transport'] = myItineraire[0]['overview_polyline']['points']
                                liste[i]['temps']['textTransport'] = myItineraire[0]['legs'][0]['duration']['text']
                                
                        
                        if myMode[j] == 'bicycling':  # quand j est 'bicycling'
                            
                         
                            liste[i]['distance']['velo'] = formatValueNum(myItineraire[0]['legs'][0]['distance']['value'],'km')
                            
                            chant['distance']['velo'].append(formatValueNum(myItineraire[0]['legs'][0]['distance']['value'],'km'))                                        
                            chant['temps']['velo'].append(round(liste[i]['distance']['velo'] / vitesseMoyenneVelo * 60))
                            chant['path']['velo'].append(myItineraire[0]['overview_polyline']['points'])
                            chant['temps']['VAE'].append(round(liste[i]['distance']['velo'] / vitesseMoyenneVAE * 60))
                            chant['etapes']['velo'].append(myItineraire[0]['legs'][0]['steps'])
                            
                            liste[i]['temps']['velo'] = round(liste[i]['distance']['velo'] / vitesseMoyenneVelo * 60)
                            liste[i]['temps']['VAE'] = round(liste[i]['distance']['velo'] / vitesseMoyenneVAE * 60)
                            liste[i]['etapes']['velo'] = myItineraire[0]['legs'][0]['steps']
                            liste[i]['path']['velo'] = myItineraire[0]['overview_polyline']['points']

         
                    else:  # si myItineraire n'existe pas... On ajoute des valeurs nuls dans la liste chantiers
                        print('erreur ditineraire')
                        if myMode[j] == 'driving':  # quand j est 'driving':
                            
                            chant['distance'].append(-1) # Sera supprimer par remove_val
                            chant['choix'].append(False)
                            chant['path'].append(None)
                            chant['temps']['voiture'].append(None)
                            chant['etapes']['voiture'].append()
                        
                        if myMode[j] == 'transit':  # quand j est 'transit':
                            chant['temps']['transport'].append(None)
                            chant['etapes']['transport'].append(None)
                            chant['path']['transport'].append(None)
                            chant['temps']['textTransport'].append(None)
                            
                        if myMode[j] == 'bicycling':  # quand j est 'bicycling':
                            chant['temps']['velo'].append(None)
                            chant['distance']['velo'].append(None)                          
                            chant['path']['velo'].append(None)
                            chant['temps']['VAE'].append(None)
                            chant['etapes']['velo'].append(None)
                            print('bicileee')
                        
                        continue
                
        if liste[i]['distance'][
            'voiture'] == 'Adresse Invalide':  # ... ça veut dire que 'adresse invalide' qui est la donnée de départ n'a pas changé
            print('[' + str(listeIndexaTraiter[i]) + '] ' + str(
                avancement) + ' % ----> INVALID ADRESS')  # donc tu me le dis!
            listeError.append(i)

        else:
            print('[' + str(listeIndexaTraiter[i]) + '] ' + str(
                avancement) + ' % ----> OK')  # sinon, c'est que tout va bien, alors tu me dit "OK"!
    
    return liste  # du coup tu me mets a jour le dictionnaire des attributs du salarie

def path (index, transport): # Renvoie le path choisi d'un index entré en paramètres
    
    if (transport == 'velo'):
        for chant in chantiers:
            if chant['choix'][index] == True:
                return chant['path']['velo'][index]
                
    if (transport == 'transport'):
        for chant in chantiers:
            if chant['choix'][index] == True:
                return chant['path']['transport'][index]
                            
    if (transport == 'voiture'):
        for chant in chantiers:
            if chant['choix'][index] == True:
                return chant['path']['voiture'][index]
                
                
def dist (index): # Renvoie la dist du chantier attribué à l'index entré en paramètres
    for chant in chantiers:
        if chant['choix'][index] == True:
            return chant['distance']['voiture'][index] 
def temps (index): # Renvoie la dist du chantier attribué à l'index entré en paramètres
    for chant in chantiers:
        if chant['choix'][index] == True:
            return chant['temps']['voiture'][index]       
def temps_velo (index): # Renvoie la dist du chantier attribué à l'index entré en paramètres
    for chant in chantiers:
        if chant['choix'][index] == True:
            return chant['temps']['velo'][index] 
def temps_transit (index): # Renvoie la dist du chantier attribué à l'index entré en paramètres
    for chant in chantiers:
        if chant['choix'][index] == True:
            return chant['temps']['transport'][index]             
 
 

def remove_sal_chant(index):
    for chant in chantiers:
        chant.remove(chant['distance'][index])
        chant.remove(chant['choix'][index])
        chant.remove(chant['temps'][index])
        chant.remove(chant['path'][index])
        chant.remove(chant['temps_transit'][index])
        chant.remove(chant['temps_velo'][index])
        chant.remove(chant['personnes'][index])
def remove_val(valeurMin, valeurMax):  # Retire des listes tout les salariés ayant un temps voiture supérieure à la valeur en paramètres
    i = 0                       
    while i < len(salarie):
        print(i)
        if salarie[i]['temps']['voiture']>valeurMax or salarie[i]['temps']['voiture'] < valeurMin:  # Si on remove pas besoin d'incrémenter i car les salariés suivants remontent d'une case dans le dictionnaire
            salarie.remove(salarie[i])
            remove_sal_chant(index)
        else : i+=1

def ecart_type():
    print(L)
    for i in range(L):
        print(i)
        if salarie[i]['info']['ecart-type'] > -1 : # On passe les ecart_type des ouvriers affectés à -1 ou moins
            ecart_type = 0
            moye = moyenne(i)
            distance = 0
            nb_chant = 0
            
            for chant in chantiers:
                if chant['plein']== False:
                    distance = chant['distance']['voiture'][i] 
                    diff = distance - moye
                    ecart_type += diff * diff # == 0 quand il n'y a qu'un seul chantier
                    nb_chant += 1 
           
            ecart_type = ecart_type/nb_chant
            ecart_type = sqrt(ecart_type) 
            salarie[i]['info']['ecart-type'] = ecart_type
def moyenne(i):  # Fonction utilisée dans l'écart-type, i = index
    nb_chant = 0    
    moyenne = 0  
    for chant in chantiers:
        if (chant['plein']== False):    #On ne prend pas en compte les chantiers pleins
            
            moyenne += chant['distance']['voiture'][i]
            nb_chant+=1
           
    moyenne = moyenne / nb_chant  
    return moyenne

def repartition ():
    index = 0

    for i in range (L):  # Répartir uniquement un % de la liste dans l'ordre des écart-type les + élevées. + intéressant que les + loins
        index = index_ec_ty()
        chant = chant_proche(index)
def index_ec_ty(): # Renvoie l'index de l'écart-type le + élevé
    ec_ty = -1
    index = -1
    for i in range(L):
        
        if ec_ty < salarie[i]['info']['ecart-type']: 
            ec_ty = salarie[i]['info']['ecart-type']
            index = i   
    if index != -1:
        salarie[index]['info']['ecart-type']= -2  # Le salarie sélectionner ne sera plus sélectionner par cette fonction 
      
    return index
def chant_proche(index):  # Renvoie le chantier le + proche du salarié entré en index

    dist = 10000
    
    for chant in chantiers:
        if chant['distance']['voiture'][index] < dist and chant['plein']==False:

            dist = chant['distance']['voiture'][index]                    
            chah = chant
    chah['places'] -=1           # Décrémente les places de 1 et si le chantier est vide recalcul les ecarts-types sans ce chantier
    if chah['places'] == 0:  # Recalcul l'écart type sans les chantiers pleins
        chah['plein'] = True
        ecart_type()
    chah['choix'][index] = True # Indique que c'est ce chantier qui a été choisi
    return chah    
    
def print_moy ():  # renvoie les moyenne de temps de trajet par chantier et le totale de celles-ci 
    moy_tot = 0
    moy_tot_bas = 0
    nb_chant = 0
    for chant in chantiers: # Pour chaque chantiers
        pond = 0
        moy_chant = 0
        moy_chant_bas = 0  # la moyenne av répartition
        nb_employ = 0
        nb_employb = 0
        for i in range(L):   # Pour chaque salarié
        
            if chant['adresse'] == salarie[i]['adresse']['affectation']: # Si l'adresse du chantier actuel est l'adresse de base du salarié actuel
                
                moy_chant_bas += salarie[i]['temps']['voiture']
                nb_employb += 1
                
            if chant['choix'][i]== True :  # Somme des chantiers choisis
                pond+=1
                moy_chant += chant['temps']['voiture'][i]
                nb_employ += 1
        try :
            
            moy_chant = moy_chant / nb_employ
            moy_chant_bas = moy_chant_bas / nb_employb
        except : print ("le chantier "+ chant['adresse']+ " a été destitué de tous ses ouvriers")
        
        print ('moyenne chantier avant répartition', chant['adresse'],'=', moy_chant_bas)
        print ('moyenne chantier :', chant['adresse'], '=', moy_chant)
        moy_tot += moy_chant * pond
        moy_tot_bas += moy_chant_bas *pond
        nb_chant += 1 * pond
        
    moy_tot = moy_tot / nb_chant
    moy_tot_bas = moy_tot_bas / nb_chant
    
    print ('moyenne total avant répartition', moy_tot_bas)
    print('moyenne total :', moy_tot)        

def columnOfSearchKey(searchKey):
    """
    Permet de trouver le numéro de la colonne dont le contenu est 'searchKey'
    Cette fonction sera utilisée dans les fonctions listBySearchKey et WriteBySearchKey
    """

    myColumnOfSearchKey = int()
    for i in range(1, maxRow + 1):  # parcours les rangées
        for j in range(1, maxCol + 1):  # et parcours les colonnes
            if maFeuilleExcel.cell(row=i, column=j).value == searchKey:  # et si tu trouves une case avec "searchKey' dedans
                myColumnOfSearchKey = j  # alors tu me donne le numéro de la colonne
    return myColumnOfSearchKey  # retourne le numéro de la colonne

def listBySearchKey(searchKey):
    """
    Récupere dans un tableau excel les valeurs par colonne selon un mot clé
    """

    liste = []
    nbCol = columnOfSearchKey(searchKey)
    
    for i in range(L):  # pour chaque rangée
        liste.append(
            str(maFeuilleExcel.cell(row=listeIndexaTraiter[i], column=nbCol).value))  # tu ajoutes à ma liste 'liste' la valeur contenue dans la case (i, colonneOfSearchkey)
    return liste  # et tu me retourne cette liste

def writeBySearchKey(liste, dictKey1, dictKey2, searchKey): #Search key en None parce que une valeur par défaut ne peut pas être suivie d'une valeur qui ne l'est pas
    """
    Écrire dans un tableau excel les valeurs dans la colonne du mot clé
    Comme le dictionnaire des attributs du salarié est de type [key1][key2],
    on donne comme argument la liste de dictionnaire (liste), et les deux clés (dictKey1 et dictKey2)
    pour accéder aux informations à écrire, puis le mot clé correspondant à la colonne excel dans laquelle les valeurs
    vont être écrite
    """
    

    for i in range(L): # pour chaque salarie
            maFeuilleExcel.cell(row = listeIndexaTraiter[i], column=columnOfSearchKey(searchKey)).value = str(liste[i][dictKey1][dictKey2])  # tu écris dans la case "rangée du salarié/colonne du searchKey", la valeur contenue dans salarie[i][attribut][sous-attributs]
            
  
def unicodeVerification(liste):
    """
    Vérification des caractères spéciaux non autorisés dans les adresses
    """

    for i in range(L):
        liste[i] = unicodedata.normalize('NFKD', liste[i]).encode('ASCII')
    return liste  # Vérification unicode

def formatValueNum(value, typeUnite):
    """
    Formater temps et distance en min et km
    le calcul d'itinéraire de l'API google renvoie des valeurs en m et s
    """
    myValue = value
    myTypeUnite = typeUnite

    if myTypeUnite == 'km':
        value = round((myValue / 1000), 1)

    elif myTypeUnite == 'mins':
        
        value = round(myValue/3600 *60)
        # value = round(myValue / 3600 * 60)

    return float(value)


def getEmissionCO2(liste):
    mode = str()

    for i in range(L):

        distanceWalking = 0
        distanceRer = 0
        distanceTram = 0
        distanceMetro = 0
        distanceBus = 0

        try:
            liste[i]['CO2']['voiture'] = round(liste[i]['distance']['voiture'] * CO2VoitureEssence, 0)
        except TypeError as errVoiture:
            liste[i]['CO2']['voiture'] = 'Indisponible'

        try:
            for j in range(0, len(liste[i]['etapes']['transport'])):
                travelMode = liste[i]['etapes']['transport'][j]['travel_mode']
                myStepDistance = liste[i]['etapes']['transport'][j]['distance']['value']

                if travelMode == 'WALKING':
                    distanceWalking += formatValueNum(myStepDistance, 'km')

                if travelMode == 'TRANSIT':
                    travelType = liste[i]['etapes']['transport'][j]['transit_details']['line']['vehicle']['type']

                    if travelType == 'TRAM':
                        distanceTram += formatValueNum(myStepDistance, 'km')
                    if travelType == 'BUS':
                        distanceBus += formatValueNum(myStepDistance, 'km')
                    if travelType == 'SUBWAY':
                        distanceMetro += formatValueNum(myStepDistance, 'km')
                    else:
                        distanceRer += formatValueNum(myStepDistance, 'km')

            liste[i]['CO2']['transport'] = round(distanceTram * CO2Tram + distanceBus * CO2Bus + distanceMetro * CO2Metro + distanceRer * CO2Rer, 0)

        except TypeError as errTransport:
            liste[i]['CO2']['transport'] = 'Indisponible'

        try:
            liste[i]['CO2']['velo'] = round(liste[i]['distance']['velo'] * CO2Velo, 0)
        except TypeError as errVelo:
            liste[i]['CO2']['velo'] = 'Indisponible'

        try:
            liste[i]['CO2']['VAE'] = round(liste[i]['distance']['voiture'] * CO2VAE, 0)
        except TypeError as errVAE:
            liste[i]['CO2']['VAE'] = 'Indisponible'

    return liste

def getElevation(liste):
    # Donne les hauteurs en m de chaque étape du trajet en vélo via Google Maps Elevation API
    gmaps = Client(API_KEY)

    for i in range(L):
        myElevations = []

        if liste[i]['solution RIDE']['velo'] == True:
            for j in range(0, len(liste[i]['etapes']['velo'])):
                stepLocation = ()
                stepLocation = (liste[i]['etapes']['velo'][j]['start_location']['lat'], liste[i]['etapes']['velo'][j]['start_location']['lng'])
                result = gmaps.elevation(stepLocation)
                myElevations.append(result[0]['elevation'])
            liste[i]['elevation']['velo'] = myElevations
        else:
            liste[i]['elevation']['velo'] = None

    return liste

def veloOuNon(liste):
    for i in range(L):
        if liste[i]['temps']['velo'] <= tempsVeloMax:
            liste[i]['solution RIDE']['velo'] = True
        else:
            liste[i]['solution RIDE']['velo'] = False
    return liste
def transportOuNon(liste):
    for i in range(L):
        try:
            if liste[i]['solution RIDE']['velo'] == False and liste[i]['solution RIDE']['VAE'] == False:
                if liste[i]['temps']['transport'] <= tempsTransportMax:
                    liste[i]['solution RIDE']['transport'] = True
                else:
                    liste[i]['solution RIDE']['transport'] = False
        except TypeError as err:
            liste[i]['solution RIDE']['transport'] = False
        else:
            liste[i]['solution RIDE']['transport'] = False

    return liste    
def voitureOuNon(liste):
    for i in range(L):

        if liste[i]['solution RIDE']['velo'] == False and liste[i]['solution RIDE']['VAE'] == False:
            if liste[i]['solution RIDE']['transport'] == False:
                liste[i]['solution RIDE']['voiture'] = True

            if liste[i]['solution RIDE']['transport'] == True and len(liste[i]['partenaires']['voisins']) != 0:
                liste[i]['solution RIDE']['voiture'] = True
                liste[i]['solution RIDE']['transport'] = False
                for item in liste[i]['partenaires']['voisins']:
                    liste[item]['solution RIDE']['voiture'] = True

        else:
            liste[i]['solution RIDE']['voiture'] = False

    return liste    
def VAEouNon(liste):
    for i in range(L):

        if liste[i]['solution RIDE']['velo'] == True:
            for j in range(1, len(liste[i]['elevation']['velo'])):
                denivele = abs(liste[i]['elevation']['velo'][j] - liste[i]['elevation']['velo'][j - 1])
                distance = liste[i]['etapes']['velo'][j]['distance']['value']
                pente = 100 * denivele / distance

                if pente >= 8:
                    liste[i]['solution RIDE']['VAE'] = True
                    liste[i]['solution RIDE']['velo'] = False
                    break
                else:
                    if j == len(liste[i]['elevation']['velo']) - 1:
                        liste[i]['solution RIDE']['VAE'] = False
                    else:
                        continue


        else:
            if liste[i]['temps']['VAE'] <= tempsVAEMax and liste[i]['solution RIDE']['velo'] == False:
                liste[i]['solution RIDE']['VAE'] = True
            else:
                liste[i]['solution RIDE']['VAE'] = False

    return liste

def checkCoworkers(liste): # Si cette fonction ne retourne rien, est-elle utile ? Vérifier l'utilisation qui en est faite
    for i in range(L):
        myCoworkers = []
        for j in range(L):
            if j != i and liste[i]['geolocation']['affectation'] == liste[j]['geolocation']['affectation']:
                myCoworkers.append(j)
        liste[i]['partenaires']['coworkers'] = myCoworkers
def checkSameAdress(liste):  # Si on retombe +sieurs x sur les mêmes adresses risque de gros décalages
    for i in range(L):
        for item in liste[i]['partenaires']['coworkers']:
            if liste[i]['geolocation']['domicile'] == liste[item]['geolocation']['domicile'] and liste[item]['geolocation']['domicile'] != None:
                lat = liste[item]['geolocation']['domicile'][0]
                lng = liste[item]['geolocation']['domicile'][1] + 0.005
                liste[item]['geolocation']['domicile'] = (lat,lng)
def checkNeighbors(liste): # Intéressant mais vérifier l'utilisation aussi
    for i in range(L):
        myNeighbors = []
        for item in liste[i]['partenaires']['coworkers']:
            distance = great_circle(liste[i]['geolocation']['domicile'], liste[item]['geolocation']['domicile']).kilometers
            if distance <= distNeighborsMax:
                myNeighbors.append(liste[item]['info']['nom']) #item ?
        liste[i]['partenaires']['voisins'] = myNeighbors
def solutionRIDE(liste):
    veloOuNon(liste)
    getElevation(liste)
    VAEouNon(liste)
    transportOuNon(liste)
    checkCoworkers(liste)
    checkNeighbors(liste)
    voitureOuNon(liste)
    for i in range(L):
        for cle, value in liste[i]['solution RIDE'].items():
            if value == True:
                print('[' + str(i + begin) + '] Solution RIDE = ' + str(cle))

def decode_polyline(polyline_str):# Doit provenir d'internet, vérifier si celles juste au dessus aussi peut-être
    index, lat, lng = 0, 0, 0
    coordinates = []
    changes = {'latitude': 0, 'longitude': 0}

    # Coordinates have variable length when encoded, so just keep
    # track of whether we've hit the end of the string. In each
    # while loop iteration, a single coordinate is decoded.
    while index < len(polyline_str):
        # Gather lat/lon changes, store them in a dictionary to apply them later
        for unit in ['latitude', 'longitude']:
            shift, result = 0, 0

            while True:
                byte = ord(polyline_str[index]) - 63
                index+=1
                result |= (byte & 0x1f) << shift
                shift += 5
                if not byte >= 0x20:
                    break

            if (result & 1):
                changes[unit] = ~(result >> 1)
            else:
                changes[unit] = (result >> 1)

        lat += changes['latitude']
        lng += changes['longitude']

        coordinates.append((lat / 100000.0, lng / 100000.0))

    return coordinates

def listeCh (adresse):  # Ajoute ou retire le chantier de la liste associée au checkbutton si l'on choche ou décoche la case du chantier
    
    if adresse in listeChant:
        listeChant.remove(adresse)
    else:
        listeChant.append(adresse)


"""def lieu(textquery,liste):
    
    
    for i in range (0,L):
        lat_dom = salarie[i]['geolocation']['domicile'][0]
        lng_dom = salarie[i]['geolocation']['domicile'][1]
        lat_afc = salarie[i]['geolocation']['affectation'][0]
        lng_afc = salarie[i]['geolocation']['affectation'][1]
        
        latitude = (lat_dom+lat_afc)/2
        longitude = (lng_dom+lng_afc)/2
      
        
    
        rayon = (salarie[i]['distance']['voiture']/2)*1000
        
        
        places_result = gmaps.places_nearby(location=(latitude,longitude), radius = rayon, open_now = False, keyword= textquery)
        for place in places_result['results']:
            lat = place['geometry']['location']['lat']
            lng = place['geometry']['location']['lng']
            print (lat)
            try:
                folium.Marker(location = (lat,lng), popup = '<b>'+str(i)+place['name']+':</b><br>'+ place['vicinity'], icon=folium.Icon(color='orange', icon_color='white', icon='building')).add_to(liste)
            except: print ('error')
    return liste"""
            
def lieu(textquery,liste): # Affiche les lieux demandés autour des points de passages similaires entre 2 trajets
    
    geolocation = [[]]*L
    for i in range(L):
        for j in range(0, len(salarie[i]['etapes']['voiture'])):
                stepLocation = ()
                #print (salarie[i]['etapes']['voiture'])
                stepLocation = (salarie[i]['etapes']['voiture'][j]['start_location']['lat'], salarie[i]['etapes']['voiture'][j]['start_location']['lng'])
                
                #print (salarie[i]['etapes']['voiture'][j][k]['start_location']['lat'])
                
                is_in = False
                for z in range(0,len(geolocation)): # Cette boucle sert à exclure les comparaisons de point de passage sur un seul et même trajet qui pourrait comporter plusieurs fois le même point de passage
                    
                    if (stepLocation in geolocation[z]) and (z != i):
                        is_in = True
                        places_result = gmaps.places_nearby(location= stepLocation, radius = 400, open_now = False, keyword= textquery)
                        for place in places_result['results']:
                            lat = place['geometry']['location']['lat']
                            lng = place['geometry']['location']['lng']

                        try:
                            folium.Marker(location = (lat,lng), popup = '<b>'+str(i)+place['name']+':</b><br>'+ place['vicinity'], icon=folium.Icon(color='orange', icon_color='white', icon='building')).add_to(liste)
                        except: print ('error')
                
                if (is_in == False):   
                    geolocation[i].append(stepLocation)
            
    print(geolocation)
    
    
    return liste
    
    
    
    
def list_passage (): # 
    
    router = Router("car")
    depart = router.findNode(lat_depart,lon_depart)
    arrivee = router.findNode(lat_arrivee,lon_arrivee)
    status, route = router.doRoute(depart, arrivee)
    if status == 'success':
        routeLatLons = list(map(router.nodeLatLon, route))
        return routeLatLons
    else: print('echec de liste')

       


        

# EXECUTE_______________________________________________________________________________________________________________



# 1) EXCEL
print('INITIALISATION-------------------------------------------------------------------------------------------------')
# 1.1) Importer le fichier Excel
monFichierExcel = oxl.load_workbook(r"C:\Users\emile.joudet\Documents\Emile Joudet\Données RIDE\FeuilleRide.xlsx")
maFeuilleExcel = monFichierExcel.worksheets[0]

# 1.2) Etendue de la feuille excel (maxRow & maxCol)
maxCol = maFeuilleExcel.max_column

listeIndexaTraiter = [i for i in range(begin, maxRow + 1)]
print('listeIndexaTraiter = ' + str(listeIndexaTraiter))

# 1.3) Preparer les listes de dictionnaires selon le nombre de salarie
L = len(listeIndexaTraiter)
for i in range(L):
    salarie.append(copy.deepcopy(attributsSalarie))

# 1.4) Remplir les infos de base  #peut-être réduire le temps de calcul en attribuant les attributs directement à 
listeAdresseDomicile = listBySearchKey('ADRESSE DOMICILE')  # la liste salarie sans passer par ces listes (si elle ne sont pas
listeNom = listBySearchKey('NOM PRENOM')        # utilisée ailleurs que dans l'affectation des attributs juste en dessous)
listeEmploi = listBySearchKey('EMPLOI')
listeAdresseAffectation = listBySearchKey('CHANTIER')

# On remplace celle-ci par celle du dessus, remplacer les fonctions associés
# listeAdresseAffectation = ['Place Pierre Jacques Dormoy, 33800 Bordeaux']*L
print(listeAdresseDomicile)


# 2) ATTRIBUTS DES SALARIES

for i in range(L): 
    salarie[i]['adresse']['domicile'] = listeAdresseDomicile[i]
    salarie[i]['adresse']['affectation'] = listeAdresseAffectation[i]
    salarie[i]['info']['nom'] = listeNom[i]
    salarie[i]['info']['emploi'] = listeEmploi[i]



### INTERFACE GRAPHIQUE -----------------------------
#fenetre = Tk()

init_chantier()


# CHECKBOX ----------------------------------------------
"""for chant in chantiers :
        
        choix_chant = Checkbutton(fenetre, text = chant['adresse'], command = partial(listeCh,chant['adresse']) )
        choix_chant.pack()
        

fenetre.mainloop()


# SUPPRESSION----------------------
for chant in chantiers:    # On supprime les chantiers pas dans la liste
    if (chant['adresse'] in listeChant) == False:
        chantiers.remove(chant)
        
i=0
while(i<len(salarie)):  # et les gens qui viennent de ces chantiers
        if (salarie[i]['adresse']['affectation'] in listeChant) == False:
            salarie.remove(salarie[i])
        else: i+= 1    
        
""" 


# 3) RIDE
print('RIDE-----------------------------------------------------------------------------------------------------------')
"""
L = len(salarie)  # On décremente L
print(L)



fenetre = Tk()

heure = Label(fenetre, text = "Heure")
heure.pack()

var_heure = IntVar()
ent_heure = Entry(fenetre,textvariable = var_heure)
ent_heure.pack()

jour = Label(fenetre, text = "Jour")
jour.pack()

var_jour = IntVar()
ent_jour = Entry(fenetre,textvariable = var_jour)
ent_jour.pack()

mois = Label(fenetre,text = "Mois")
mois.pack()

var_mois = IntVar()
ent_mois = Entry(fenetre,textvariable = var_mois)
ent_mois.pack()


button_ok = Button(fenetre, command = lambda :getItineraire(salarie,int(ent_heure.get()),int(ent_jour.get()),int(ent_mois.get())))
button_ok.pack()

fenetre.mainloop()
"""
getItineraire(salarie)


#getItineraire(salarie)
"""
fenetre = Tk()

temps_lim_min = Label(fenetre, text = "Temps Limite min")
temps_lim_min.pack()

var_entre = IntVar()
entre = Entry(fenetre,textvariable = var_entre)
entre.pack()

temps_lim = Label(fenetre, text = "Temps Limite max")
temps_lim.pack()


var_entree = IntVar()
entree = Entry(fenetre, textvariable = var_entree)
entree.pack()

button_ok = Button(fenetre, command = lambda :remove_val(int(entre.get()),int(entree.get())))
button_ok.pack()

fenetre.mainloop()

"""
"""L = len(salarie)"""

for chant in chantiers:
    print("ici")
    print(chant['adresse'])

"""ecart_type()
repartition()
print_moy()

for i in range(L):     # on remplace dans salarie les données avant répartition par les résultats après répartition pour les afficher
    for chant in chantiers:
        if chant['choix'][i] == True :
            salarie[i]['temps']['voiture'] = chant['temps']['voiture'][i]
            salarie[i]['temps']['transit'] = chant['temps']['transport'][i]
            salarie[i]['etapes']['transport'] = chant['etapes']['transport'][i]
            salarie[i]['etapes']['velo'] = chant['etapes']['velo'][i]
            salarie[i]['etapes']['voiture']=chant['etapes']['voiture']
            salarie[i]['temps']['textTransport'] = chant['temps']['textTransport'][i]
            salarie[i]['path']['voiture'] = chant['path']['voiture'][i]
            salarie[i]['path']['velo'] = chant['path']['velo'][i]
            salarie[i]['path']['transport'] = chant['path']['transport'][i]
            salarie[i]['temps']['velo'] = chant['temps']['velo'][i]
            salarie[i]['etapes']['transport'] = chant ['etapes']['transport'][i]
            salarie[i]['etapes']['velo'] = chant ['etapes']['velo'][i]
            salarie[i]['adresse']['affectation'] = chant['adresse']
            salarie[i]['distance']['voiture'] = chant['distance']['voiture'][i]
            salarie[i]['distance']['velo'] = chant['distance']['velo'][i]
         
            
            
getEmissionCO2(salarie)
checkCoworkers(salarie)
checkSameAdress(salarie)
checkNeighbors(salarie)
solutionRIDE(salarie)
"""
#liste_pass = list_passage(47.6005615, 1.311846,48.5529093, 2.2137127)


#liste_pass = list_passage(salarie[1]['geolocation']['domicile'][0],salarie[1]['geolocation']['domicile'][1],salarie[1]['geolocation']['affectation'][0],salarie[1]['geolocation']['affectation'][1])
#print(liste_pass)


# 6) EXPORTER LES DONNEES DANS UN FICHIER EXCEL
print('EXPORT DATA----------------------------------------------------------------------------------------------------')

# GENERAL

writeBySearchKey(salarie, 'geolocation', 'domicile', 'GEOLOCATION DOMICILE')
writeBySearchKey(salarie, 'adresse', 'affectation', 'CHANTIER')
writeBySearchKey(salarie, 'info','nom', 'NOM PRENOM')
writeBySearchKey(salarie,'adresse', 'domicile','ADRESSE DOMICILE')

#writeBySearchKey(chantiers,'distance',None,'GEOLOCATION CHANTIER') 


# VOITURE
writeBySearchKey(salarie, 'distance', 'voiture', 'Distance domicile-chantier') 
writeBySearchKey(salarie, 'temps', 'voiture', 'Temps voiture')
writeBySearchKey(salarie, 'CO2', 'voiture', 'Emission voiture')
# TRANSPORT
writeBySearchKey(salarie, 'temps', 'transport', 'Temps transport')
writeBySearchKey(salarie, 'CO2', 'transport', 'Emission transport')
# VELO
writeBySearchKey(salarie, 'temps', 'velo', 'Temps vélo')
writeBySearchKey(salarie, 'CO2', 'velo', 'Emission vélo')
writeBySearchKey(salarie, 'partenaires', 'voisins', 'Voisins')

# On écrit les noms des chantiers en tete de liste

# SAUVEGARDER

monFichierExcel.save(r"C:\Users\emile.joudet\Documents\Emile Joudet\Données RIDE\FeuilleRide_Result.xlsx") # Cette fonction n'est pas encore dans le main je crois


# 7) CREER LA CARTE INTERACTIVE
print('CREATION DE LA CARTE INTERACTIVE-------------------------------------------------------------------------------')
myMap = folium.Map(location = (44.858189, -0.574842), tiles=None)

# LAYER VOITURE
# légende
description = 'TRAJETS EN VOITURE:' 

# création du layer voiture et des sous-layers
layerVoiture = folium.FeatureGroup(name=description, overlay=True, control=True)
layer_voit_15 = plugins.FeatureGroupSubGroup(layerVoiture, name = '<font color=green> <big><b>&#9632;</b></big></font> : De 0 à 15 minutes', show = True, overlay = True)
layer_voit_30 = plugins.FeatureGroupSubGroup(layerVoiture, name = '<font color=orange> <big><b>&#9632;</b></big></font> : De 15 à 30 minutes', show = True, overlay = True)
layer_voit_45 = plugins.FeatureGroupSubGroup(layerVoiture, name = '<font color=red> <big><b>&#9632;</b></big></font> : De 30 à 45 minutes', show = True, overlay = True)
layer_voit_60 = plugins.FeatureGroupSubGroup(layerVoiture, name = '<font color=darkred> <big><b>&#9632;</b></big></font> : Supérieur à 45 min', show = True, overlay = True)

# polyligne et markers selon temps de trajet

for i in range(L):
  colMarker = str()
  colPline = str()

  if salarie[i]['adresse']['affectation'] == 'Noisy-Legrand':
        colMarker = 'green'
        colPline = colorSet['green']
        
  elif salarie[i]['adresse']['affectation'] == 'Villejuif':
        colMarker = 'orange'
        colPline = colorSet['orange']
        
  elif salarie[i]['adresse']['affectation'] == 'CNIT LA DEFENSE':
        colMarker = 'red'
        colPline = 'red'

  """  # if temps(i) is not None:
  if salarie[i]['temps']['voiture'] < 15:
        colMarker = 'green'
        colPline = colorSet['green']
        
  elif 15 < salarie[i]['temps']['voiture'] < 30:
        colMarker = 'orange'
        colPline = colorSet['orange']
        
  elif 30 < salarie[i]['temps']['voiture'] < 45:
        colMarker = 'red'
        colPline = 'red'
        
  elif 45 < salarie[i]['temps']['voiture'] < 60:
        colMarker = 'darkred'
        colPline = 'darkred'
        
  elif 60 < salarie[i]['temps']['voiture'] :
        colMarker = 'darkred'
        colPline = 'darkred'   """

  popup = '<b>%s:</b><br>' \
            '<li>Centre de profit : <b>%s</b></li>' \
            '<li><b>%s</b> pour rejoindre le chantier <b>%s</b></li>' \
            '<li>Trajet de <b>%s</b> kms</li>' \
            '<li><b>%s</b> g de CO2 rejetés</li>' %(salarie[i]['info']['nom'],
                                              salarie[i]['info']['emploi'],
                                              salarie[i]['adresse']['affectation'],
                                              round(salarie[i]['temps']['voiture'],0),
                                              salarie[i]['distance']['voiture'],
                                              salarie[i]['CO2']['voiture'])
    
  try:  # Remplissage des sous-layers
        if salarie[i]['temps']['voiture'] < 15 : 
            folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_voit_15)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='black', weight=7, opacity=1).add_to(layer_voit_15) #Poids inversement proportionelle à la distance
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='gray', weight=6, opacity=1).add_to(layer_voit_15)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_voit_15)
        
        if 15 < salarie[i]['temps']['voiture'] < 30 :
            folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_voit_30)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='black', weight=7, opacity=1).add_to(layer_voit_30) #Poids inversement proportionelle à la distance
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='gray', weight=6, opacity=1).add_to(layer_voit_30)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_voit_30)
            
        if 30 < salarie[i]['temps']['voiture'] < 45 :
            folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_voit_45)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='black', weight=7, opacity=1).add_to(layer_voit_45) #Poids inversement proportionelle à la distance
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='gray', weight=6, opacity=1).add_to(layer_voit_45)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_voit_45)
            
        if 45 < salarie[i]['temps']['voiture']:
            folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_voit_60)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='black', weight=7, opacity=1).add_to(layer_voit_60) #Poids inversement proportionelle à la distance
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color='gray', weight=6, opacity=1).add_to(layer_voit_60)
            folium.PolyLine(decode_polyline(salarie[i]['path']['voiture']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_voit_60)    
     
  except : continue

top_left = (min([s['geolocation']['domicile'][0] for s in salarie]), min([s['geolocation']['domicile'][1] for s in salarie]))
bottom_right = (max([s['geolocation']['domicile'][0] for s in salarie]), max([s['geolocation']['domicile'][1] for s in salarie]))




# marker chantier
# marquer tous les chantier en faisant une boucle qui marque si la variable contenant la geo du chantier a change
# marker chantier
getGeo()
for chant in chantiers: 
    print(chant['adresse'])
    try:
        folium.Marker(location = chant['geolocation'], popup = '<b>CHANTIER:</b><br>'+ chant['adresse'], icon=folium.Icon(color='blue', icon_color='white', icon='building')).add_to(layerVoiture)
    except: print ('error')

lieu('porte',layerVoiture) # Fonction de recherche avec mot clé
    
#folium.Marker(location=salarie[5]['geolocation']['affectation'], popup = '<b>CHANTIER:</b><br>Place Pierre Jacques Dormoy 33800 BORDEAUX', icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerVoiture)
#folium.Marker(location=(44.826693, -0.564527), popup = '<b>CHANTIER:</b><br>Place Pierre Jacques Dormoy 33800 BORDEAUX', icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerVoiture)

# ajout du layer voiture à myMap

layerVoiture.add_to(myMap)
layer_voit_15.add_to(myMap)
layer_voit_30.add_to(myMap)
layer_voit_45.add_to(myMap)
layer_voit_60.add_to(myMap)
mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
mapTile.add_to(layerVoiture)




# LAYER TRANSPORT
# légende
reperage = []
for i in range(L):
    if salarie[i]['etapes']['transport'] is not None:
        for step in salarie[i]['etapes']['transport']:
            if step['travel_mode'] == 'TRANSIT':
                try:
                    col = step['transit_details']['line']['color']
                except KeyError as error:
                    col = 'black'
                try:  #KeyError name 
                 type = '%s %s' % (step['transit_details']['line']['vehicle']['name'], step['transit_details']['line']['name'])
                 reperage.append('<br><font color=%s><big><b>&mdash;</b></big></font> : %s' % (col, type))
                except : print('error name')
    else:
        continue
reperage = set(reperage)
description = 'TRAJETS EN TRANSPORTS EN COMMUN:' 
              
              
# création du layer transport
layerTransport = folium.FeatureGroup(name=description, overlay=True, control=True)
layer_trans_15 = plugins.FeatureGroupSubGroup(layerTransport, name = '<font color=green> <big><b>&#9632;</b></big></font> :  De 0 à 15 minutes', show = True, overlay = True)
layer_trans_30 = plugins.FeatureGroupSubGroup(layerTransport, name = '<font color=orange> <big><b>&#9632;</b></big></font> :  De 15 à 30 minutes', show = True, overlay = True)
layer_trans_45 = plugins.FeatureGroupSubGroup(layerTransport, name = '<font color=red> <big><b>&#9632;</b></big></font> :  De 30 à 45 minutes', show = True, overlay = True)
layer_trans_60 = plugins.FeatureGroupSubGroup(layerTransport, name = '<font color=darkred><big><b>&#9632;</b></big></font>: Supérieur à 45 min  <br><big><b>&#9632;</b></big></font>: Zone non désservie <br><br> <i>Repérage des transports en commun :</i>'+''.join(reperage), show = True, overlay = True)




# polyligne et markers selon temps de trajet
for i in range(L):
    colMarker = str()
    if salarie[i]['temps']['transport'] is not None:
        if 0 <= salarie[i]['temps']['transport'] <= 15:
            colMarker = 'green'
        elif 15 < salarie[i]['temps']['transport'] <= 30:
            colMarker = 'orange'
        elif 30 < salarie[i]['temps']['transport'] <= 45:
            colMarker = 'red'
        elif 45 < salarie[i]['temps']['transport'] <= 60:
            colMarker = 'darkred'
        elif salarie[i]['temps']['transport'] > 60:
            colMarker = 'darkpurple'
    else:
        colMarker = 'black'

#instruction de transport :
    instructions = []
    if salarie[i]['etapes']['transport'] is not None:
        for step in salarie[i]['etapes']['transport']:
            if step['travel_mode'] == 'TRANSIT' :
                try:
                 instructions.append('<li>Prendre <b>%s %s</b> de <i>%s</i> à <i>%s</i> (%s arrêts, %s)</li>' % (step['transit_details']['line']['vehicle']['name'],
                                                                                                                step['transit_details']['line']['name'],
                                                                                                                step['transit_details']['departure_stop']['name'],
                                                                                                                step['transit_details']['arrival_stop']['name'],
                                                                                                                step['transit_details']['num_stops'],
                                                                                                                step['duration']['text']))
                                                                                                                
                except : continue
                                                                                                                
            if step['travel_mode'] == 'WALKING':
                instructions.append('<li>%s (%s)</li>' % (step['html_instructions'], step['duration']['text']))

        popup = '<b>%s:</b><br>' \
                '<li>Centre de profit : <b>%s</b></li>' \
                '<li><b>%s</b> pour rejoindre le chantier</li>' \
                '<li>Instructions de trajet :</li>' \
                '<ul>%s</ul>' \
                '<li><b>%s</b> g de CO2 rejetés</li>' % (salarie[i]['info']['nom'],
                                                   salarie[i]['info']['emploi'],
                                                   salarie[i]['temps']['textTransport'],
                                                   ''.join(instructions),
                                                   salarie[i]['CO2']['transport'])
    else:
        popup = '<b>%s:</b><br>' \
                '<li>Centre de profit : <b>%s</b></li>' \
                '<li><b>ZONE NON DÉSSERVIE PAR LES TRANSPORTS EN COMMUN</b></li>' % (salarie[i]['info']['nom'], salarie[i]['info']['emploi'],)
    
    if salarie[i]['temps']['transport'] != None:
        if salarie[i]['temps']['transport'] < 15 :
        
            if (salarie[i]['geolocation']['domicile'] is not None):
                folium.Marker(location=salarie[i]['geolocation']['domicile'], popup=popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant'] == 'Encadrement Tvx' else 'wrench'),prefix='fa')).add_to(layer_trans_15)

            if salarie[i]['etapes']['transport'] is not None:
                for j in range(0, len(salarie[i]['etapes']['transport'])):
                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'TRANSIT':
                        try:
                            colPline = salarie[i]['etapes']['transport'][j]['transit_details']['line']['color']
                        except KeyError as error :
                            colPline = 'black'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_15)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_15)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_15)


                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'WALKING':
                        colPline = '#FFFFFF'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_15)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_15)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_15)
            else:
                continue
        
        if 15 < salarie[i]['temps']['transport'] < 30 :
        
            if (salarie[i]['geolocation']['domicile'] is not None):
                folium.Marker(location=salarie[i]['geolocation']['domicile'], popup=popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant'] == 'Encadrement Tvx' else 'wrench'),prefix='fa')).add_to(layer_trans_30)

            if salarie[i]['etapes']['transport'] is not None:
                for j in range(0, len(salarie[i]['etapes']['transport'])):
                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'TRANSIT':
                        try:
                            colPline = salarie[i]['etapes']['transport'][j]['transit_details']['line']['color']
                        except KeyError as error :
                            colPline = 'black'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_30)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_30)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_30)


                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'WALKING':
                        colPline = '#FFFFFF'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_30)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_30)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_30)
            else:
                continue
        
        if 30 < salarie[i]['temps']['transport'] < 45 :
        
            if (salarie[i]['geolocation']['domicile'] is not None):
                folium.Marker(location=salarie[i]['geolocation']['domicile'], popup=popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant'] == 'Encadrement Tvx' else 'wrench'),prefix='fa')).add_to(layer_trans_45)

            if salarie[i]['etapes']['transport'] is not None:
                for j in range(0, len(salarie[i]['etapes']['transport'])):
                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'TRANSIT':
                        try:
                            colPline = salarie[i]['etapes']['transport'][j]['transit_details']['line']['color']
                        except KeyError as error :
                            colPline = 'black'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_45)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_45)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_45)


                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'WALKING':
                        colPline = '#FFFFFF'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_45)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_45)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_45)
            else:
                continue
    
        if 45 < salarie[i]['temps']['transport'] :
        
            if (salarie[i]['geolocation']['domicile'] is not None):
                folium.Marker(location=salarie[i]['geolocation']['domicile'], popup=popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant'] == 'Encadrement Tvx' else 'wrench'),prefix='fa')).add_to(layer_trans_60)

            if salarie[i]['etapes']['transport'] is not None:
                for j in range(0, len(salarie[i]['etapes']['transport'])):
                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'TRANSIT':
                        try:
                            colPline = salarie[i]['etapes']['transport'][j]['transit_details']['line']['color']
                        except KeyError as error :
                            colPline = 'black'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_60)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_60)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_60)


                    if salarie[i]['etapes']['transport'][j]['travel_mode'] == 'WALKING':
                        colPline = '#FFFFFF'

                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layer_trans_60)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layer_trans_60)
                        folium.PolyLine(decode_polyline(salarie[i]['etapes']['transport'][j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layer_trans_60)
            else:
                continue
    
    
    
    
# Marker chantiers
for chant in chantiers: 
    try:
        folium.Marker(location = chant['geolocation'], popup = '<b>CHANTIER:</b><br>'+ chant['adresse'], icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerTransport)
    except: print ('error')

#ajout du layer transport à myMap
layerTransport.add_to(myMap)
layer_trans_15.add_to(myMap)
layer_trans_30.add_to(myMap)
layer_trans_45.add_to(myMap)
layer_trans_60.add_to(myMap)
mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
mapTile.add_to(layerTransport)


# LAYER VELO
# légende
# création du layer vélo


description = 'TRAJETS EN VELO:' 
             
              
#LAYER VELO et sous layers 
              
layerVelo = folium.FeatureGroup(name=description, overlay=True, control=True, show = False)
layerVoiture = folium.FeatureGroup(name=description, overlay=True, control=True)
layer_velo_15 = plugins.FeatureGroupSubGroup(layerVelo, name = '<font color=green> <big><b>&#9632;</b></big></font>  : De 0 à 15 minutes', show = True, overlay = True)
layer_velo_30 = plugins.FeatureGroupSubGroup(layerVelo, name = '<font color=orange> <big><b>&#9632;</b></big></font>  : De 15 à 30 minutes', show = True, overlay = True)
layer_velo_45 = plugins.FeatureGroupSubGroup(layerVelo, name = '<font color=red> <big><b>&#9632;</b></big></font>  : De 30 à 45 minutes', show = True, overlay = True)
layer_velo_60 = plugins.FeatureGroupSubGroup(layerVelo, name = '<font color=darkred> <big><b>&#9632;</b></big></font>  : Supérieur à 45 min <br>', show = True, overlay = True)

# polyligne et markers selon temps de trajet

for i in range(L):
  colMarker = str()
  colPline = str()
    
  if salarie[i]['temps']['velo'] is not None:
    if 0 < salarie[i]['temps']['velo']< 15 :
        colMarker = 'green'
        colPline = colorSet['green']
        
    elif 15 < salarie[i]['temps']['velo']< 30 :
        colMarker = 'orange'
        colPline = colorSet['orange']
        
    elif 30 < salarie[i]['temps']['velo']< 45 :
        colMarker = 'red'
        colPline = 'red'
        
    elif 45 < salarie[i]['temps']['velo']< 60 :
        colMarker = 'darkred'
        colPline = 'darkred'
        
    elif 60 < salarie[i]['temps']['velo'] :
        colMarker = 'darkpurple'
        colPline = colorSet['purple']

  popup = '<b>%s:</b><br>' \
            '<li>Centre de profit : <b>%s</b></li>' \
            '<li><b>%s</b> pour rejoindre le chantier <b>%s</b></li>' \
            '<li>Trajet de <b>%s</b> kms</li>' \
            '<li><b>%s</b> g de CO2 rejetés</li>' %(salarie[i]['info']['nom'],
                                              salarie[i]['info']['emploi'],
                                              salarie[i]['adresse']['affectation'],
                                              round(salarie[i]['temps']['velo'],0),
                                              salarie[i]['distance']['velo'],
                                              salarie[i]['CO2']['velo'])
  
  
  try:
      
     if salarie[i]['temps']['velo']< 15:
        folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_velo_15)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='black', weight=7, opacity=1).add_to(layer_velo_15) #Poids inversement proportionelle à la distance
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='gray', weight=6, opacity=1).add_to(layer_velo_15)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_velo_15)
     
     if 15 < salarie[i]['temps']['velo']< 30:
        folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_velo_30)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='black', weight=7, opacity=1).add_to(layer_velo_30) #Poids inversement proportionelle à la distance
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='gray', weight=6, opacity=1).add_to(layer_velo_30)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_velo_30)
     
     if 30 < salarie[i]['temps']['velo']< 45:
        folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_velo_45)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='black', weight=7, opacity=1).add_to(layer_velo_45) #Poids inversement proportionelle à la distance
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='gray', weight=6, opacity=1).add_to(layer_velo_45)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_velo_45)
        
     if 45 < salarie[i]['temps']['velo']:
        folium.Marker(location=salarie[i]['geolocation']['domicile'], popup = popup, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase' if salarie[i]['info']['encadrant']=='Encadrement Tvx' else 'wrench'), prefix='fa')).add_to(layer_velo_60)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='black', weight=7, opacity=1).add_to(layer_velo_60) #Poids inversement proportionelle à la distance
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color='gray', weight=6, opacity=1).add_to(layer_velo_60)
        folium.PolyLine(decode_polyline(salarie[i]['path']['velo']), color=colPline, weight=4.5, opacity=(0.8)).add_to(layer_velo_60)
  except : continue

top_left = (min([s['geolocation']['domicile'][0] for s in salarie]), min([s['geolocation']['domicile'][1] for s in salarie]))
bottom_right = (max([s['geolocation']['domicile'][0] for s in salarie]), max([s['geolocation']['domicile'][1] for s in salarie]))




# marker chantier
# marquer tous les chantier en faisant une boucle qui marque si la variable contenant la geo du chantier a change
# marker chantier
getGeo()
for chant in chantiers: 
    print(chant['adresse'])
    try:
        folium.Marker(location = chant['geolocation'], popup = '<b>CHANTIER:</b><br>'+ chant['adresse'], icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerVelo)
    except: print ('error')

    
#folium.Marker(location=salarie[5]['geolocation']['affectation'], popup = '<b>CHANTIER:</b><br>Place Pierre Jacques Dormoy 33800 BORDEAUX', icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerVoiture)
#folium.Marker(location=(44.826693, -0.564527), popup = '<b>CHANTIER:</b><br>Place Pierre Jacques Dormoy 33800 BORDEAUX', icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerVoiture)

# ajout du layer voiture à myMap
layerVelo.add_to(myMap)
layer_velo_15.add_to(myMap)
layer_velo_30.add_to(myMap)
layer_velo_45.add_to(myMap)
layer_velo_60.add_to(myMap)
mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
mapTile.add_to(layerVelo)


"""# CHECKBOX ----------------------------------------------
# Checkbox des calques

fenetre = Tk()
       
choix_calque_voiture = Checkbutton(fenetre, text = 'Voiture', command = lambda: list_Map(layerVoiture))
choix_calque_voiture.pack()

choix_calque_transport = Checkbutton(fenetre, text = 'Transport',command = lambda: list_Map(layerTransport))
choix_calque_transport.pack()

button_ok = Button(fenetre, text = 'OK', command = lambda : create_map())
button_ok.pack()    


        

fenetre.mainloop()"""



# AUTRES SETTINGS DE LA MAP ET SAUVEGARDE
layerPanel = folium.LayerControl(position='topleft',collapsed=False,autoZIndex=True)
layerPanel.add_to(myMap)
logoISCpath = 'https://www.construction21.org/france/data/sources/users/3957/logo-vinci-construction-fr.jpg'
logoSOGEApath = 'https://aragodesete.fr/wp-content/uploads/2018/04/Log_SOGEA_Sud_Batiment_C_R.png'
logoISC = folium.plugins.FloatImage(image=logoISCpath, bottom=1, left=12)
logoSOGEA = folium.plugins.FloatImage(image=logoSOGEApath, bottom=0.5, left=16.5)
#logoISC.add_to(myMap) On les enlèves our l'instant soucis avec scale
#logoSOGEA.add_to(myMap)
myMap.fit_bounds(bounds=[top_left,bottom_right])
myMap.save(r"C:\Users\emile.joudet\Documents\Emile Joudet\Données RIDE\Save.html")

# Interface Utilisateur

class Interface(Frame):
    
    def __init__(self, fenetre, **kwargs):
        Frame.__init__(self, fenetre, width=768, height=576, **kwargs)
        self.pack(fill=BOTH)
        
        self.temps_trajet = 0
        self.temps_trajet_trans = 0
        self.temps_trajet_velo = 0
        self.distance_trajet = 0
        self.distance_trajet_velo = 0
        self.text_transport = ""
        self.text_voiture = ""
        
        self.poids = 0
        
        self.path_trajet = " " 
        self.path_trajet_trans= " "
        self.path_trajet_velo = " "
        
        
        self.etapes_transport = []
        self.etapes_velo = []
        
        self.lab = Label(self, text = "Poids en kg (optionnel)")
        self.lab.pack(fill = BOTH)
        
        self.ent_poids = Entry(self, textvariable = IntVar())
        self.ent_poids.pack(fill = BOTH)
        
        
        self.Ad_dom = Label(self, text = "Adresse Domicile")
        self.Ad_dom.pack(fill=BOTH)
        self.geo_dom = (0,0)

        self.ent_dom = Entry(self)
        self.ent_dom.pack(fill=BOTH)
        
        self.Ad_prof = Label(self, text = "Adresse professionelle")
        self.Ad_prof.pack(fill=BOTH)
        self.geo_prof = (0,0)
        
        self.ent_prof = Entry(self)
        self.ent_prof.pack(fill=BOTH)


        self.button_ok = Button(self, text = 'Rechercher', command = self.push)
        self.button_ok.pack(fill=BOTH)
        
    
    def CO2_transport (self):
        
        distanceWalking = 0
        distanceRer = 0
        distanceTram = 0
        distanceMetro = 0
        distanceBus = 0
        CO2_transport = 0
        
        for j in range(0, len(self.etapes_transport)):
                travelMode = self.etapes_transport[j]['travel_mode']
                myStepDistance = self.etapes_transport[j]['distance']['value']

                if travelMode == 'WALKING':
                    distanceWalking += formatValueNum(myStepDistance, 'km')

                if travelMode == 'TRANSIT':
                    travelType = self.etapes_transport[j]['transit_details']['line']['vehicle']['type']

                    if travelType == 'TRAM':
                        distanceTram += formatValueNum(myStepDistance, 'km')
                    if travelType == 'BUS':
                        distanceBus += formatValueNum(myStepDistance, 'km')
                    if travelType == 'SUBWAY':
                        distanceMetro += formatValueNum(myStepDistance, 'km')
                    else:
                        distanceRer += formatValueNum(myStepDistance, 'km')
                        
        CO2_transport = round(distanceTram * CO2Tram + distanceBus * CO2Bus + distanceMetro * CO2Metro + distanceRer * CO2Rer, 0)
        return CO2_transport
    
    def Instru(self):
        
            instructions = []
            if self.etapes_transport is not None:
                for step in self.etapes_transport:
                    if step['travel_mode'] == 'TRANSIT' :
            
                        try:
                            instructions.append('<li>Prendre <b>%s %s</b> de <i>%s</i> à <i>%s</i> (%s arrêts; %s)</li><br />; ' % (step['transit_details']['line']['vehicle']['name'],
                                                                                                                step['transit_details']['line']['name'],
                                                                                                                step['transit_details']['departure_stop']['name'],
                                                                                                                step['transit_details']['arrival_stop']['name'],
                                                                                                                step['transit_details']['num_stops'],
                                                                                                                step['duration']['text']))
                                                                                                                
                        
                        except: 
                            instructions.append('<li>Prendre <b>%s</b> de <i>%s</i> à <i>%s</i> (%s arrêts; %s)</li><br />; ' % (step['transit_details']['line']['vehicle']['name'],
                                                                                                            
                                                                                                                step['transit_details']['departure_stop']['name'],
                                                                                                                step['transit_details']['arrival_stop']['name'],
                                                                                                                step['transit_details']['num_stops'],
                                                                                                                step['duration']['text']))
                                                                                                                
                                                                                                                
                    if step['travel_mode'] == 'WALKING':
                        instructions.append('<li>%s (%s)</li><br /> ;' % (step['html_instructions'], step['duration']['text']))
        
            soup = BeautifulSoup(str(instructions))
            instru = soup.get_text().replace(';','\n')
        
        
            top=Toplevel(self)
            top.geometry('500x100')
            lab=Label(top, text = instru)
            lab.pack() 
        
        
    
    def push(self):
        
        self.getItin ()
        self.Display_Toplevel()
        
    def getItin (self):
            
            if self.ent_poids != 0 :
                try : 
                   self.poids = int(self.ent_poids.get())
                except: 
                    print("La valeur entrée dans poids est invalide")
            
            now = datetime.now()
            adresse_dom = self.ent_dom.get()
            adresse_prof = self.ent_prof.get()
            try :
                myItineraire = gmaps.directions(adresse_dom, adresse_prof, mode=str('driving'), region='FR', departure_time=now, language='fr', traffic_model='best_guess')
                print('succes')
                self.temps_trajet = round(myItineraire[0]['legs'][0]['duration_in_traffic']['value']*0.016666667,0)
                self.distance_trajet = formatValueNum(myItineraire[0]['legs'][0]['distance']['value'],'km')
               
                self.path_trajet = myItineraire[0]['overview_polyline']['points']
                self.geo_dom = (myItineraire[0]['legs'][0]['start_location']['lat'], myItineraire[0]['legs'][0]['start_location']['lng'])  # tu ajoutes un tupple (lat, lng) dans le dictionnaire aux clés [géolocation][domicile] que l'on trouve dans le dictionnaire de résultats de google
                self.geo_prof = (myItineraire[0]['legs'][0]['end_location']['lat'],myItineraire[0]['legs'][0]['end_location']['lng'])
                self.text_voiture = myItineraire[0]['legs'][0]['duration']['text']
                
            except : 
                print ("Adresse(s) Invalide(s)")
                
            try : 
                myItineraire = gmaps.directions(adresse_dom, adresse_prof, mode=str('transit'), region='FR', departure_time=now, language='fr', traffic_model='best_guess')    
                self.etapes_transport = myItineraire[0]['legs'][0]['steps']  
                self.path_trajet_trans = myItineraire[0]['overview_polyline']['points']  
                self.temps_trajet_trans = myItineraire[0]['legs'][0]['duration']['value']*0.016666667
                self.text_transport = myItineraire[0]['legs'][0]['duration']['text']
                        
            except: 
                print("itinéraire en transport Indisponible")
                
            try : 
                myItineraire = gmaps.directions(adresse_dom, adresse_prof, mode=str('bicycling'), region='FR', departure_time=now, language='fr', traffic_model='best_guess')    
                
                self.distance_trajet_velo = formatValueNum(myItineraire[0]['legs'][0]['distance']['value'],'km')
                self.etapes_velo = myItineraire[0]['legs'][0]['steps']
                self.temps_trajet_velo = round(self.distance_trajet_velo / vitesseMoyenneVelo * 60)
                
                self.path_trajet_velo = myItineraire[0]['overview_polyline']['points']
                   
                        
            except: 
                print("itinéraire en transport Indisponible")    
    
    def Display_Toplevel(self):
        top=Toplevel(self)
        top.geometry('520x200')
        
        lab=Label(top, text = " Votre trajet de : "+ str(self.ent_dom.get())  + "\n à : "+ str(self.ent_prof.get()) + " \n \n \n En voiture : dure " + str(self.temps_trajet) + " min et représente " + str(self.distance_trajet) + " km. Vous émettez " + str(round(self.distance_trajet*CO2VoitureEssence,0)) + "g de CO2 " \
                                                       " \n En transport : dure " + str(round(self.temps_trajet_trans,0)) + " min. Vous émettez " + str(self.CO2_transport()) + "g de CO2 " \
                                                       " \n En vélo : dure " + str(self.temps_trajet_velo) + " min et représente " + str(self.distance_trajet_velo) + " km.. Vous émettez " + str(round(self.distance_trajet_velo * CO2Velo, 0)) + "g de CO2 et vous perdez "+ str(self.poids * self.temps_trajet_velo * 0.1 )+ " kcal." )
                                                                                                                        
        lab.grid(row = 1, column = 2)  
        
        button_instructions = Button(top, text = 'Instructions de Transport', command = self.Instru)
        button_instructions.grid(row = 3, column = 2)
        
        button_instructions = Button(top, text = 'Dénivelé Vélo', command = self.denivele)
        button_instructions.grid(row = 4, column = 2)
        
        button_carte = Button(top, text = 'Afficher Carte', command = self.carte)
        button_carte.grid(row = 5, column = 2)
        
        
    def carte (self):
       
        myMap2 = folium.Map(location = (44.858189, -0.574842), tiles=None)
        
        
        
        # Layer Voiture
        
        layerVoit = folium.FeatureGroup(name= 'Trajet en Voiture', overlay=True, control=True)
        
        pop = '<li><b>%s</b> pour rejoindre le chantier <b>%s</b> min </li>' \
            '<li>Trajet de <b>%s</b> kms</li>' \
            '<li><b>%s</b> g de CO2 rejetés</li>' %(self.ent_dom.get(),
                                              round(self.temps_trajet,0),
                                              self.distance_trajet,
                                              round(self.distance_trajet * CO2VoitureEssence, 0))
        
        
        
        
        folium.Marker(location= self.geo_dom, popup = pop, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase'), prefix='fa')).add_to(layerVoit)
        folium.Marker(location = self.geo_prof, popup = '<b>CHANTIER:</b><br>'+ self.ent_prof.get(), icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerVoit)
        
        folium.PolyLine(decode_polyline(self.path_trajet), color='black', weight=7, opacity=1).add_to(layerVoit) #Poids inversement proportionelle à la distance
        folium.PolyLine(decode_polyline(self.path_trajet), color='gray', weight=6, opacity=1).add_to(layerVoit)
        folium.PolyLine(decode_polyline(self.path_trajet), color='blue', weight=4.5, opacity=(0.8)).add_to(layerVoit)
        
        layerVoit.add_to(myMap2)
        mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
        mapTile.add_to(layerVoit)
        
        
        #Layer velo
        
        layerVelo = folium.FeatureGroup(name= 'Trajet en Velo', overlay=True, control=True)
        
        
        pop = '<li><b>%s</b> pour rejoindre le chantier <b>%s</b> min </li>' \
            '<li>Trajet de <b>%s</b> kms</li>' \
            '<li><b>%s</b> g de CO2 rejetés</li>' %(self.ent_dom.get(),
                                              round(self.temps_trajet_velo,0),
                                              self.distance_trajet_velo,
                                              round(self.distance_trajet_velo * CO2Velo, 0))
        
        folium.Marker(location= self.geo_dom, popup = pop, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase'), prefix='fa')).add_to(layerVelo)
        folium.Marker(location = self.geo_prof, popup = '<b>CHANTIER:</b><br>'+ self.ent_prof.get(), icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerVelo)
        
        folium.PolyLine(decode_polyline(self.path_trajet_velo), color='black', weight=7, opacity=1).add_to(layerVelo) #Poids inversement proportionelle à la distance
        folium.PolyLine(decode_polyline(self.path_trajet_velo), color='gray', weight=6, opacity=1).add_to(layerVelo)
        folium.PolyLine(decode_polyline(self.path_trajet_velo), color='green', weight=4.5, opacity=(0.8)).add_to(layerVelo)
        
        layerVelo.add_to(myMap2)
        mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
        mapTile.add_to(layerVelo)
        
        # Layer Transport
        
    
        # légende
        reper = []
        for i in range(L):
            if self.etapes_transport is not None:
                for step in self.etapes_transport:
                    if step['travel_mode'] == 'TRANSIT':
                        try:
                            col = step['transit_details']['line']['color']
                        except KeyError as error:
                            col = 'black'
                        try:  #KeyError name 
                            type = '%s %s' % (step['transit_details']['line']['vehicle']['name'], step['transit_details']['line']['name'])
                            reper.append('<br><font color=%s><big><b>&mdash;</b></big></font> : %s' % (col, type))
                        except : print('error name')
            else:
                continue
        reper = set(reper)
      
        pop = ''
        
        instructions = []
        if self.etapes_transport is not " ":
            for step in self.etapes_transport:
                if step['travel_mode'] == 'TRANSIT' :
                    try:
                        instructions.append('<li>Prendre <b>%s %s</b> de <i>%s</i> à <i>%s</i> (%s arrêts, %s)</li>' % (step['transit_details']['line']['vehicle']['name'],
                                                                                                                step['transit_details']['line']['name'],
                                                                                                                step['transit_details']['departure_stop']['name'],
                                                                                                                step['transit_details']['arrival_stop']['name'],
                                                                                                                step['transit_details']['num_stops'],
                                                                                                                step['duration']['text']))
                                                                                                                
                    except:
                        instructions.append('<li>Prendre <b>%s</b> de <i>%s</i> à <i>%s</i> (%s arrêts, %s)</li>' % (step['transit_details']['line']['vehicle']['name'],
                                                                                                               
                                                                                                                step['transit_details']['departure_stop']['name'],
                                                                                                                step['transit_details']['arrival_stop']['name'],
                                                                                                                step['transit_details']['num_stops'],
                                                                                                                step['duration']['text']))
                        
                if step['travel_mode'] == 'WALKING':
                    instructions.append('<li>%s (%s)</li>' % (step['html_instructions'], step['duration']['text']))
                    
            pop ='<li><b>%s</b> pour rejoindre le chantier</li>' \
                    '<li>Instructions de trajet :</li>' \
                    '<ul>%s</ul>' \
                    '<li><b>%s</b> g de CO2 rejetés</li>' % (self.text_transport,
                                                   ''.join(instructions),
                                                   str(self.CO2_transport()))
        else:
            pop = '<li><b>ZONE NON DÉSSERVIE PAR LES TRANSPORTS EN COMMUN</b></li>' 
        
        
        layerTrans = folium.FeatureGroup(name= 'Trajet en Transport </br> <i>Repérage des transports en commun :</i>'+''.join(reper), overlay=True, control=True, show = False)
        
        folium.Marker(location= self.geo_dom, popup = pop, icon=folium.Icon(color=colMarker, icon_color='white', icon=('briefcase'), prefix='fa')).add_to(layerTrans)
        folium.Marker(location = self.geo_prof, popup = '<b>CHANTIER:</b><br>'+ self.ent_prof.get(), icon=folium.Icon(color='blue', icon_color='white', icon='building', prefix= 'fa')).add_to(layerTrans)
        
        for j in range(0, len(self.etapes_transport)):
                    if self.etapes_transport[j]['travel_mode'] == 'TRANSIT':
                        try:
                            colPline = self.etapes_transport[j]['transit_details']['line']['color']
                        except KeyError as error :
                            colPline = 'black'

                        folium.PolyLine(decode_polyline(self.etapes_transport[j]['polyline']['points']), color='black', weight=7,opacity=1).add_to(layerTrans)
                        folium.PolyLine(decode_polyline(self.etapes_transport[j]['polyline']['points']), color='gray', weight=6,opacity=1).add_to(layerTrans)
                        folium.PolyLine(decode_polyline(self.etapes_transport[j]['polyline']['points']), color=colPline, weight=4.5,opacity=(0.8)).add_to(layerTrans)
        
        layerTrans.add_to(myMap2)
        mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
        mapTile.add_to(layerTrans)
        
        
        # Parc relais _ Aire de Covoit - 
        
        LayerParc = folium.FeatureGroup(name= 'Parc Relais', overlay=True, control=True, show = False)
        LayerCovoit = folium.FeatureGroup(name= 'Aire de covoiturage', overlay=True, control=True, show = False)
        
        lat_dom = self.geo_dom[0]
        lng_dom = self.geo_dom[1]
        lat_afc = self.geo_prof[0]
        lng_afc = self.geo_prof[1]
        
        latitude = (lat_dom+lat_afc)/2
        longitude = (lng_dom+lng_afc)/2
      
        
    
        rayon = (self.distance_trajet/2)*1000
        
        # Recherche des parcs relais
        places_result = gmaps.places_nearby(location=(latitude,longitude), radius = rayon, open_now = False, keyword= 'parking relais', type = 'parking')
        for place in places_result['results']:
            lat = place['geometry']['location']['lat']
            lng = place['geometry']['location']['lng']
            try:
                folium.Marker(location = (lat,lng), popup = '<b>'+place['name']+':</b><br>'+ place['vicinity'], icon=folium.Icon(color='green', icon_color='white', icon='building')).add_to(LayerParc)
            except: print ('error')
        
        # Recherche des aires de covoiturage
        places_result = gmaps.places_nearby(location=(latitude,longitude), radius = rayon, open_now = False, keyword= 'Aire de covoiturage')
        for place in places_result['results']:
            lat = place['geometry']['location']['lat']
            lng = place['geometry']['location']['lng']
            try:
                folium.Marker(location = (lat,lng), popup = '<b>'+place['name']+':</b><br>'+ place['vicinity'], icon=folium.Icon(color='blue', icon_color='white', icon='building')).add_to(LayerCovoit)
            except: print ('error')
        
        
        
        LayerParc.add_to(myMap2)
        mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
        mapTile.add_to(LayerParc)
        
        LayerCovoit.add_to(myMap2)
        mapTile = folium.TileLayer(tiles='OpenStreetMap') #StamenToner
        mapTile.add_to(LayerCovoit)
        
        
        
        
        
        layerPanel = folium.LayerControl(position='topleft',collapsed=False,autoZIndex=True)
        layerPanel.add_to(myMap2)
        
        
        
        myMap2.fit_bounds(bounds=[top_left,bottom_right])
        myMap2.save(r"C:\Users\emile.joudet\Documents\Emile Joudet\Données RIDE\Save2.html")
        
        style_statement = '<style>.leaflet-control{color:#00FF00}</style>'
        #myMap2.get_root().html.add_child(folium.Element(style_statement))
        html_string = myMap2.get_root().render()
        
        print(self.etapes_transport)
        
        f = open("carte.html","w")
        f.write(html_string)
        f.close()
        
        webbrowser.open('carte.html')
        
      
        
    def denivele (self):  # Donne les hauteurs en m de chaque étape du trajet en vélo via Google Maps Elevation API
    

        myElevations = []
        index = []
        dist = 0
        
        
        for j in range(0, len(self.etapes_velo)):
            stepLocation = ()
            stepLocation = (self.etapes_velo[j]['start_location']['lat'], self.etapes_velo[j]['start_location']['lng'])
            dist += self.etapes_velo[j]['distance']['value']
            result = gmaps.elevation(stepLocation)
            myElevations.append(result[0]['elevation'])
            index.append(dist)
        
        
        blue_patch = mpatches.Patch(color='b', label=' pente < 4% ')
        yellow_patch = mpatches.Patch(color='yellow', label=' 4% < pente < 8% ')
        red_patch = mpatches.Patch(color='r', label=' pente > 8% ')
     
        plt.legend(handles=[blue_patch, yellow_patch, red_patch])

        
        for j in range(1, len(myElevations)):
            denivele = abs(myElevations[j] - myElevations[j - 1])
            distance = self.etapes_velo[j]['distance']['value']
            pente = 100 * denivele / distance

            if 4 < pente < 8 or -8< pente <-4:
                
                plt.plot([index[j-1],index[j]], [myElevations[j-1],myElevations[j]], 'yellow')
            
            if pente > 8 or pente < -8 :
                
                plt.plot([index[j-1],index[j]], [myElevations[j-1],myElevations[j]], 'r')
                
            else:
                
                plt.plot([index[j-1],index[j]], [myElevations[j-1],myElevations[j]], 'b')
                
        plt.xlabel('distance trajet (m)')
        plt.ylabel('denivele (m)')
        
        plt.show()
        

        
        
fenetre = Tk()
interface = Interface(fenetre)

interface.mainloop()
interface.destroy()
